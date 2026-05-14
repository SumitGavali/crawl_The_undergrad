"""
KnowledgeOS — Complete Backend
Real crawler + embeddings + search, all triggered from the UI.
One button → reads files → extracts text → creates embeddings → saves DB → search works.
"""

import json
import logging
import os
import pickle
import platform
import subprocess
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
from dotenv import load_dotenv
from fastapi import FastAPI, Query, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

# ── Load .env ──────────────────────────────────────────────────────────────────
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("knowledgeos")

# ── Config from .env ───────────────────────────────────────────────────────────
MODEL_NAME          = os.getenv("MODEL_NAME", "all-MiniLM-L6-v2")
TOP_K               = int(os.getenv("TOP_K", "5"))
CHUNK_SIZE          = int(os.getenv("CHUNK_SIZE", "400"))
CHUNK_OVERLAP       = int(os.getenv("CHUNK_OVERLAP", "50"))
API_HOST            = os.getenv("API_HOST", "0.0.0.0")
API_PORT            = int(os.getenv("API_PORT", "8000"))
DATA_DIR            = os.path.join(os.path.dirname(__file__), os.getenv("DATA_DIR", "./data"))
RELEVANCE_THRESHOLD = float(os.getenv("RELEVANCE_THRESHOLD", "0.15"))
BATCH_SIZE          = int(os.getenv("BATCH_SIZE", "64"))
QDRANT_URL          = os.getenv("QDRANT_URL", "http://localhost:6333")
QDRANT_COLLECTION   = os.getenv("QDRANT_COLLECTION", "knowledgeos")
VECTOR_DB_PATH      = os.path.join(DATA_DIR, "vectors.pkl")
METADATA_PATH       = os.path.join(DATA_DIR, "metadata.json")
SUPPORTED_EXT       = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".txt", ".csv"}
MIN_CHUNK_WORDS     = 8

# ── Network Drive (SMB) config ─────────────────────────────────────────────────
NETWORK_DRIVE_ENABLED  = os.getenv("NETWORK_DRIVE_ENABLED", "false").lower() == "true"
NETWORK_DRIVE_HOST     = os.getenv("NETWORK_DRIVE_HOST", "")
NETWORK_DRIVE_SHARE    = os.getenv("NETWORK_DRIVE_SHARE", "")
NETWORK_DRIVE_USERNAME = os.getenv("NETWORK_DRIVE_USERNAME", "")
NETWORK_DRIVE_PASSWORD = os.getenv("NETWORK_DRIVE_PASSWORD", "")
NETWORK_DRIVE_DOMAIN   = os.getenv("NETWORK_DRIVE_DOMAIN", "")
NETWORK_DRIVE_CACHE_DIR = os.path.join(DATA_DIR, "network_drive_cache")

# ── SharePoint (Graph API) config ───────────────────────────────────────────────
SHAREPOINT_ENABLED       = os.getenv("SHAREPOINT_ENABLED", "false").lower() == "true"
SHAREPOINT_TENANT_ID     = os.getenv("SHAREPOINT_TENANT_ID", "")
SHAREPOINT_CLIENT_ID     = os.getenv("SHAREPOINT_CLIENT_ID", "")
SHAREPOINT_CLIENT_SECRET = os.getenv("SHAREPOINT_CLIENT_SECRET", "")
SHAREPOINT_SITE_URL      = os.getenv("SHAREPOINT_SITE_URL", "")
SHAREPOINT_CACHE_DIR     = os.path.join(DATA_DIR, "sharepoint_cache")

# ── Scheduler config ───────────────────────────────────────────────────────────────
REINDEX_INTERVAL_HOURS = int(os.getenv("REINDEX_INTERVAL_HOURS", "24"))
REINDEX_ON_STARTUP     = os.getenv("REINDEX_ON_STARTUP", "false").lower() == "true"
# ───────────────────────────────────────────────────────────────────────────────

os.makedirs(DATA_DIR, exist_ok=True)


# ── Text extractors ───────────────────────────────────────────────────────────

def extract_pdf(path: str) -> list[dict]:
    try:
        import pdfplumber
        results = []
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                if text.strip():
                    results.append({"text": text.strip(), "page": i + 1})
        return results
    except Exception as e:
        logger.error("PDF extraction failed for %s: %s", path, e)
        return [{"text": f"PDF error: {e}", "page": 1}]


def extract_docx(path: str) -> list[dict]:
    try:
        from docx import Document
        doc = Document(path)
        full = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        words = full.split()
        results = []
        for i in range(0, max(1, len(words)), CHUNK_SIZE):
            chunk = " ".join(words[i:i + CHUNK_SIZE])
            if chunk.strip():
                results.append({"text": chunk, "page": i // CHUNK_SIZE + 1})
        return results
    except Exception as e:
        logger.error("DOCX extraction failed for %s: %s", path, e)
        return [{"text": f"DOCX error: {e}", "page": 1}]


def extract_xlsx(path: str) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        results = []
        for i, name in enumerate(wb.sheetnames):
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = "  |  ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    rows.append(row_text)
            if rows:
                results.append({"text": f"[Sheet: {name}]\n" + "\n".join(rows), "page": i + 1})
        return results
    except Exception as e:
        logger.error("XLSX extraction failed for %s: %s", path, e)
        return [{"text": f"XLSX error: {e}", "page": 1}]


def extract_txt(path: str) -> list[dict]:
    try:
        text = Path(path).read_text(encoding="utf-8", errors="ignore")
        words = text.split()
        results = []
        for i in range(0, max(1, len(words)), CHUNK_SIZE):
            chunk = " ".join(words[i:i + CHUNK_SIZE])
            if chunk.strip():
                results.append({"text": chunk, "page": i // CHUNK_SIZE + 1})
        return results
    except Exception as e:
        logger.error("TXT extraction failed for %s: %s", path, e)
        return [{"text": f"TXT error: {e}", "page": 1}]


def extract_text(path: str) -> list[dict]:
    ext = Path(path).suffix.lower()
    if ext == ".pdf":
        return extract_pdf(path)
    elif ext in (".docx", ".doc"):
        return extract_docx(path)
    elif ext in (".xlsx", ".xls"):
        return extract_xlsx(path)
    else:
        return extract_txt(path)


def make_chunks(pages: list[dict], file_path: str, file_type: str) -> list[dict]:
    chunks = []
    step = CHUNK_SIZE - CHUNK_OVERLAP
    for page in pages:
        words = page["text"].split()
        for i in range(0, len(words), step):
            w = words[i:i + CHUNK_SIZE]
            if len(w) < MIN_CHUNK_WORDS:
                continue
            chunks.append({
                "file_path":   file_path,
                "file_type":   file_type,
                "page":        page["page"],
                "chunk_index": len(chunks),
                "text":        " ".join(w),
                "source":      "local",  # Default source for local files
            })
    return chunks


# ── Path validation ────────────────────────────────────────────────────────────

def validate_folder_path(folder: str) -> tuple[bool, str]:
    """Reject non-absolute paths and directory-traversal attempts."""
    if not folder or not folder.strip():
        return False, "Folder path cannot be empty."
    resolved = os.path.realpath(os.path.abspath(folder))
    if not os.path.isabs(folder):
        return False, f"Path must be absolute (got relative: {folder!r}). Provide a full path like /home/user/docs."
    # Block traversal: resolved path must start with the originally intended parent
    if ".." in folder:
        return False, "Directory traversal ('..') is not allowed in the path."
    if not os.path.isdir(resolved):
        return False, f"Folder not found: {resolved}"
    return True, resolved


# ── SearchService (replaces global mutable state) ─────────────────────────────

from cachetools import TTLCache
from db.vector_store import QdrantVectorStore


class SearchService:
    """Holds the embedding model and Qdrant vector store. One instance per app."""

    def __init__(self, model_name: str = MODEL_NAME):
        self.vector_store: Optional[QdrantVectorStore] = None
        self.model = None
        self.model_loaded = False
        self._query_cache: TTLCache = TTLCache(maxsize=1000, ttl=3600)

        try:
            from sentence_transformers import SentenceTransformer
            logger.info("Loading embedding model '%s' …", model_name)
            self.model = SentenceTransformer(model_name)
            self.model_loaded = True
            logger.info("✓ Embedding model loaded")
        except Exception as e:
            logger.warning("Could not load SentenceTransformer: %s", e)

        # Initialize Qdrant vector store
        try:
            self.vector_store = QdrantVectorStore(QDRANT_URL, QDRANT_COLLECTION)
            logger.info("✓ Connected to Qdrant at %s", QDRANT_URL)
        except Exception as e:
            logger.error("Failed to connect to Qdrant: %s", e)
            raise

    # -- migration ------------------------------------------------------------

    def migrate_from_pickle(self) -> None:
        """Migrate existing pickle data to Qdrant, then delete old files."""
        if not os.path.exists(VECTOR_DB_PATH):
            logger.info("No pickle file found, skipping migration")
            return

        logger.info("Found existing pickle data, migrating to Qdrant...")
        
        try:
            # Load old data
            with open(VECTOR_DB_PATH, "rb") as f:
                vecs = pickle.load(f)
            if isinstance(vecs, list):
                vecs = np.array(vecs, dtype=np.float32)
            
            metadata = []
            if os.path.exists(METADATA_PATH):
                with open(METADATA_PATH, "r", encoding="utf-8") as f:
                    metadata = json.load(f)
            
            if len(vecs) != len(metadata):
                logger.warning("Vector/metadata mismatch: %d vectors, %d metadata", len(vecs), len(metadata))
                return
            
            # Add source field if missing
            for chunk in metadata:
                if "source" not in chunk:
                    chunk["source"] = "local"
            
            # Upsert to Qdrant
            self.vector_store.upsert(metadata, vecs)
            logger.info("✓ Migrated %d vectors to Qdrant", len(vecs))
            
            # Delete old files
            os.remove(VECTOR_DB_PATH)
            logger.info("✓ Deleted %s", VECTOR_DB_PATH)
            if os.path.exists(METADATA_PATH):
                os.remove(METADATA_PATH)
                logger.info("✓ Deleted %s", METADATA_PATH)
            
        except Exception as e:
            logger.error("Migration failed: %s", e)
            raise

    # -- search ---------------------------------------------------------------

    def search(self, query: str, sources: Optional[list[str]] = None) -> dict:
        """Encode *query* and return ranked results. Results are TTL-cached.

        Parameters
        ----------
        query   : Search string (minimum 2 chars).
        sources : Optional list of source tags to restrict results to
                  (e.g. ``["local", "sharepoint"]``).  ``None`` means all sources.
        """
        cache_key = query.strip().lower()
        if sources:
            cache_key += "|".join(sorted(sources))
        if cache_key in self._query_cache:
            logger.info("Cache hit for query: %s", cache_key)
            return self._query_cache[cache_key]

        start = time.time()

        if self.vector_store is None:
            return {
                "query": query, "results": [], "total": 0,
                "elapsed_ms": 0, "mode": "no_index",
                "message": "Vector store not initialized."
            }

        # Check if we have any data
        stats = self.vector_store.get_stats()
        if stats.get("total_points", 0) == 0:
            result = {
                "query": query, "results": [], "total": 0,
                "elapsed_ms": 0, "mode": "no_index",
                "message": "No index found. Go to Live Crawl tab and index your folder first."
            }
            return result

        if not self.model_loaded:
            return {"error": "Embedding model not loaded.", "results": [], "mode": "error"}

        # Encode query and search
        qvec = self.model.encode([query])[0].astype(np.float32)
        hits = self.vector_store.search(qvec, top_k=TOP_K, sources=sources or None)

        results = []
        for hit in hits:
            if hit["score"] < RELEVANCE_THRESHOLD:
                continue
            payload = hit["payload"]
            results.append({
                "file":   payload.get("file_path", ""),
                "chunk":  payload.get("text", ""),
                "score":  float(round(hit["score"], 4)),
                "page":   payload.get("page", 1),
                "type":   payload.get("file_type", ""),
                "source": payload.get("source", "local"),
            })

        response = {
            "query": query, "results": results, "total": len(results),
            "elapsed_ms": round((time.time() - start) * 1000), "mode": "live",
            "sources": sources,
        }
        self._query_cache[cache_key] = response
        logger.info("Search for '%s' → %d results in %dms", query, len(results), response["elapsed_ms"])
        return response

    # -- update ---------------------------------------------------------------

    def update(self, new_vectors: np.ndarray, new_metadata: list[dict]) -> None:
        """Add new vectors and metadata to Qdrant."""
        self.vector_store.upsert(new_metadata, new_vectors)
        self._query_cache.clear()
        logger.info("Index updated: %d vectors added", len(new_vectors))


# ── Background crawl task tracking ─────────────────────────────────────────────

crawl_tasks: dict[str, dict] = {}   # task_id → { status, progress, … }

# Stores last-run timestamps for the connectors/status endpoint
_connector_last_indexed: dict[str, str] = {}   # connector_name → ISO-8601 timestamp


def background_crawl(task_id: str, folder: str, search_svc: SearchService) -> None:
    """Run the full crawl→chunk→embed pipeline in the background."""
    task = crawl_tasks[task_id]
    try:
        task["status"] = "running"
        task["message"] = f"Scanning {folder}"
        logger.info("[task:%s] Starting crawl of %s", task_id, folder)

        all_files = []
        for root, _, files in os.walk(folder):
            for fname in files:
                if Path(fname).suffix.lower() in SUPPORTED_EXT:
                    all_files.append(os.path.join(root, fname))

        if not all_files:
            task["status"] = "error"
            task["message"] = "No supported files found."
            logger.warning("[task:%s] No files found in %s", task_id, folder)
            return

        task["total_files"] = len(all_files)
        task["message"] = f"Found {len(all_files)} files"
        logger.info("[task:%s] Found %d files", task_id, len(all_files))

        # Extract + chunk
        all_chunks: list[dict] = []
        for i, fpath in enumerate(all_files, 1):
            ext = Path(fpath).suffix.upper().strip(".")
            pages = extract_text(fpath)
            chunks = make_chunks(pages, fpath, ext)
            all_chunks.extend(chunks)
            task["files_processed"] = i
            task["total_chunks"] = len(all_chunks)
            task["message"] = f"Processed {i}/{len(all_files)} files"
            logger.info("[task:%s] Processed %s → %d chunks", task_id, fpath, len(chunks))

        if not all_chunks:
            task["status"] = "error"
            task["message"] = "Could not extract any text from the files."
            return

        if not search_svc.model_loaded:
            task["status"] = "error"
            task["message"] = "Embedding model not loaded."
            return

        # Embed
        task["message"] = f"Creating embeddings for {len(all_chunks)} chunks …"
        texts = [c["text"] for c in all_chunks]
        all_vecs: list = []

        for b in range(0, len(texts), BATCH_SIZE):
            batch = texts[b:b + BATCH_SIZE]
            vecs = search_svc.model.encode(batch, show_progress_bar=False)
            all_vecs.extend(vecs)
            pct = min(100, round(((b + len(batch)) / len(texts)) * 100))
            task["embedding_pct"] = pct
            task["message"] = f"Embedding {b + len(batch)}/{len(texts)} ({pct}%)"

        vectors_arr = np.array(all_vecs, dtype=np.float32)

        # Save & hot-reload
        search_svc.update(vectors_arr, all_chunks)

        task["status"] = "complete"
        task["message"] = f"Done — {len(all_files)} files, {len(all_chunks)} chunks. Ready to search."
        logger.info("[task:%s] Crawl complete: %d files, %d chunks", task_id, len(all_files), len(all_chunks))

    except Exception as exc:
        task["status"] = "error"
        task["message"] = f"Crawl failed: {exc}"
        logger.exception("[task:%s] Crawl failed", task_id)


# ── Real Crawl SSE (kept for the existing frontend) ───────────────────────────

def real_crawl_generator(folder: str, search_svc: SearchService):
    def evt(event, data):
        return f"event: {event}\ndata: {json.dumps(data)}\n\n"

    folder = os.path.abspath(folder)
    yield evt("start", {"message": f"Starting crawl → {folder}", "folder": folder})
    time.sleep(0.15)

    if not os.path.exists(folder):
        yield evt("error", {"message": f"Folder not found: {folder}"})
        logger.error("Crawl target not found: %s", folder)
        return

    all_files = []
    for root, _, files in os.walk(folder):
        for f in files:
            if Path(f).suffix.lower() in SUPPORTED_EXT:
                all_files.append(os.path.join(root, f))

    if not all_files:
        yield evt("error", {"message": "No supported files found. Supported types: PDF, DOCX, XLSX, TXT, CSV"})
        return

    yield evt("info", {"message": f"Found {len(all_files)} files to process", "total": len(all_files), "demo": False})
    time.sleep(0.15)

    all_chunks: list[dict] = []
    indexed = 0

    for i, fpath in enumerate(all_files):
        ext = Path(fpath).suffix.upper().strip(".")
        yield evt("reading",    {"file": fpath, "type": ext, "index": i+1, "total": len(all_files), "status": "reading"})
        pages = extract_text(fpath)
        yield evt("extracting", {"file": fpath, "type": ext, "pages": len(pages), "index": i+1, "status": "extracting"})
        chunks = make_chunks(pages, fpath, ext)
        yield evt("indexing",   {"file": fpath, "chunks": len(chunks), "type": ext, "index": i+1, "status": "indexing"})
        all_chunks.extend(chunks)
        indexed += 1
        yield evt("file_done",  {
            "file": fpath, "chunks": len(chunks), "type": ext,
            "indexed": indexed, "total": len(all_files), "total_chunks": len(all_chunks)
        })
        logger.info("Crawl-SSE: processed %s → %d chunks", fpath, len(chunks))

    if not all_chunks:
        yield evt("error", {"message": "Could not extract any text from the files."})
        return

    if not search_svc.model_loaded:
        yield evt("error", {"message": "Embedding model not loaded. Run: pip install sentence-transformers"})
        return

    yield evt("embedding_start", {"message": f"Creating embeddings for {len(all_chunks)} chunks...", "total_chunks": len(all_chunks)})

    texts    = [c["text"] for c in all_chunks]
    all_vecs: list = []

    for b in range(0, len(texts), BATCH_SIZE):
        batch = texts[b:b + BATCH_SIZE]
        vecs  = search_svc.model.encode(batch, show_progress_bar=False)
        all_vecs.extend(vecs)
        pct = min(100, round(((b + len(batch)) / len(texts)) * 100))
        yield evt("embedding_progress", {"done": b + len(batch), "total": len(texts), "pct": pct})

    vectors_arr = np.array(all_vecs, dtype=np.float32)

    yield evt("saving", {"message": "Saving index to disk..."})
    search_svc.update(vectors_arr, all_chunks)

    yield evt("complete", {
        "total_files":  len(all_files),
        "total_chunks": len(all_chunks),
        "message": f"Index complete — {len(all_files)} files, {len(all_chunks)} chunks. Ready to search."
    })
    logger.info("Crawl-SSE complete: %d files, %d chunks", len(all_files), len(all_chunks))


# ── App lifespan ───────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: init SearchService, DocumentAgent; shutdown cleanly."""
    import threading
    from scheduler.agent import DocumentAgent

    # ── SearchService ──────────────────────────────────────────────────────
    svc = SearchService(model_name=MODEL_NAME)
    svc.migrate_from_pickle()
    app.state.search_service = svc

    # ── DocumentAgent ──────────────────────────────────────────────────────
    agent = DocumentAgent(
        vector_store=svc.vector_store,
        search_service=svc,
        last_indexed_registry=_connector_last_indexed,
        interval_hours=REINDEX_INTERVAL_HOURS,
    )
    agent.start()
    app.state.agent = agent

    if REINDEX_ON_STARTUP:
        logger.info("REINDEX_ON_STARTUP=true — triggering initial indexing run")
        threading.Thread(
            target=agent.run_all_sources,
            name="startup-reindex",
            daemon=True,
        ).start()

    logger.info("KnowledgeOS started — host=%s port=%s", API_HOST, API_PORT)
    yield
    # ── Shutdown ───────────────────────────────────────────────────────────
    agent.stop()
    logger.info("KnowledgeOS shutting down")


app = FastAPI(title="KnowledgeOS API", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _svc() -> SearchService:
    return app.state.search_service

def _agent():
    return app.state.agent


# ── SSE crawl endpoint (existing frontend uses this) ───────────────────────────

@app.get("/api/crawl/stream")
def crawl_stream(folder: str = ""):
    return StreamingResponse(
        real_crawl_generator(folder or "./demo_docs", _svc()),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


# ── Async crawl endpoint (new — background task) ──────────────────────────────

@app.post("/api/crawl")
async def start_crawl(folder: str, background_tasks: BackgroundTasks):
    """Kick off a crawl in the background. Returns immediately with a task_id."""
    valid, msg = validate_folder_path(folder)
    if not valid:
        logger.warning("Rejected crawl path: %s — %s", folder, msg)
        return {"error": msg}

    task_id = str(uuid.uuid4())
    crawl_tasks[task_id] = {
        "status": "queued",
        "message": "Queued",
        "total_files": 0,
        "files_processed": 0,
        "total_chunks": 0,
        "embedding_pct": 0,
    }
    background_tasks.add_task(background_crawl, task_id, msg, _svc())
    logger.info("Crawl task queued: %s → %s", task_id, msg)
    return {"task_id": task_id, "status": "queued"}


@app.get("/api/crawl/status/{task_id}")
def crawl_status(task_id: str):
    """Poll progress of a background crawl."""
    task = crawl_tasks.get(task_id)
    if task is None:
        return {"error": f"Unknown task_id: {task_id}"}
    return {"task_id": task_id, **task}


# ── Network Drive indexing ─────────────────────────────────────────────────────

def background_index_network_drive(task_id: str, search_svc: SearchService) -> None:
    """
    Background task: connect → list → download → embed → upsert for the
    configured SMB share.  Updates crawl_tasks[task_id] throughout.
    """
    task = crawl_tasks[task_id]
    try:
        task["status"] = "running"
        task["message"] = f"Connecting to //{NETWORK_DRIVE_HOST}/{NETWORK_DRIVE_SHARE} …"
        logger.info("[task:%s] Starting network-drive indexing", task_id)

        from connectors.network_drive import NetworkDriveConnector
        from connectors.indexing_service import index_network_drive

        connector = NetworkDriveConnector(
            host=NETWORK_DRIVE_HOST,
            share=NETWORK_DRIVE_SHARE,
            username=NETWORK_DRIVE_USERNAME,
            password=NETWORK_DRIVE_PASSWORD,
            domain=NETWORK_DRIVE_DOMAIN,
        )

        task["message"] = "Running indexing pipeline …"
        summary = index_network_drive(
            search_service=search_svc,
            vector_store=search_svc.vector_store,
            connector=connector,
            cache_dir=NETWORK_DRIVE_CACHE_DIR,
            extract_fn=extract_text,
            make_chunks_fn=make_chunks,
            batch_size=BATCH_SIZE,
        )

        if summary.get("error"):
            task["status"] = "error"
            task["message"] = summary["message"]
            logger.error("[task:%s] Network-drive indexing failed: %s", task_id, summary["message"])
        else:
            task["status"] = "complete"
            task["message"] = summary["message"]
            task["summary"] = summary
            _connector_last_indexed["network_drive"] = datetime.utcnow().isoformat()
            logger.info("[task:%s] Network-drive indexing complete: %s", task_id, summary)

    except Exception as exc:
        task["status"] = "error"
        task["message"] = f"Network drive indexing failed: {exc}"
        logger.exception("[task:%s] Unhandled error in network-drive indexing", task_id)


@app.post("/api/index/network-drive")
async def index_network_drive_endpoint(background_tasks: BackgroundTasks):
    """
    Trigger a background SMB-share indexing run.

    - Returns 400 if ``NETWORK_DRIVE_ENABLED`` is not true in .env.
    - Returns 400 if any required credential env var is missing.
    - Otherwise queues the job and returns a ``task_id`` for polling via
      ``GET /api/crawl/status/{task_id}``.
    """
    from fastapi import HTTPException

    if not NETWORK_DRIVE_ENABLED:
        raise HTTPException(
            status_code=400,
            detail=(
                "Network drive connector is disabled. "
                "Set NETWORK_DRIVE_ENABLED=true in .env to enable it."
            ),
        )

    missing = [v for v, val in [
        ("NETWORK_DRIVE_HOST",     NETWORK_DRIVE_HOST),
        ("NETWORK_DRIVE_SHARE",    NETWORK_DRIVE_SHARE),
        ("NETWORK_DRIVE_USERNAME", NETWORK_DRIVE_USERNAME),
        ("NETWORK_DRIVE_PASSWORD", NETWORK_DRIVE_PASSWORD),
    ] if not val]

    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required env vars: {', '.join(missing)}",
        )

    task_id = str(uuid.uuid4())
    crawl_tasks[task_id] = {
        "status":      "queued",
        "message":     "Queued — waiting to start",
        "indexed":     0,
        "skipped":     0,
        "failed":      0,
        "total":       0,
        "total_chunks": 0,
    }
    background_tasks.add_task(background_index_network_drive, task_id, _svc())
    logger.info("Network-drive indexing task queued: %s", task_id)
    return {"task_id": task_id, "status": "queued"}


@app.get("/api/connectors/status")
def connectors_status():
    """
    Return enabled/disabled status and last-indexed timestamp for all connectors.

    Queries Qdrant for point counts per source so callers can see how many
    chunks each connector has contributed.
    """
    svc = _svc()

    def _source_stats(source: str) -> dict:
        """Return chunk / file counts for a given source value."""
        try:
            from qdrant_client.models import Filter, FieldCondition, MatchValue
            scroll_result = svc.vector_store.client.scroll(
                collection_name=svc.vector_store.collection_name,
                scroll_filter=Filter(
                    must=[FieldCondition(key="source", match=MatchValue(value=source))]
                ),
                limit=10000,
                with_payload=True,
                with_vectors=False,
            )
            unique_files: set = set()
            for point in scroll_result[0]:
                unique_files.add(point.payload.get("file_path", ""))
            return {"chunks": len(scroll_result[0]), "files": len(unique_files)}
        except Exception as exc:
            logger.warning("Could not fetch stats for source=%s: %s", source, exc)
            return {"chunks": 0, "files": 0, "error": str(exc)}

    local_stats = _source_stats("local")
    nd_stats    = _source_stats("network_drive") if NETWORK_DRIVE_ENABLED else {}
    sp_stats    = _source_stats("sharepoint")    if SHAREPOINT_ENABLED    else {}

    return {
        "connectors": {
            "local": {
                "enabled":      True,
                "last_indexed": _connector_last_indexed.get("local"),
                **local_stats,
            },
            "network_drive": {
                "enabled":      NETWORK_DRIVE_ENABLED,
                "host":         NETWORK_DRIVE_HOST  or None,
                "share":        NETWORK_DRIVE_SHARE or None,
                "last_indexed": _connector_last_indexed.get("network_drive"),
                **(nd_stats if NETWORK_DRIVE_ENABLED else {"chunks": None, "files": None}),
            },
            "sharepoint": {
                "enabled":      SHAREPOINT_ENABLED,
                "site_url":     SHAREPOINT_SITE_URL or None,
                "last_indexed": _connector_last_indexed.get("sharepoint"),
                **(sp_stats if SHAREPOINT_ENABLED else {"chunks": None, "files": None}),
            },
        }
    }


# ── SharePoint indexing ──────────────────────────────────────────────────────────

def background_index_sharepoint(task_id: str, search_svc: SearchService) -> None:
    """
    Background task: authenticate → list → download → embed → upsert for
    the configured SharePoint site.  Updates crawl_tasks[task_id] throughout.
    """
    task = crawl_tasks[task_id]
    try:
        task["status"]  = "running"
        task["message"] = f"Connecting to SharePoint: {SHAREPOINT_SITE_URL} …"
        logger.info("[task:%s] Starting SharePoint indexing", task_id)

        from connectors.sharepoint import SharePointConnector
        from connectors.indexing_service import index_sharepoint

        connector = SharePointConnector(
            tenant_id=SHAREPOINT_TENANT_ID,
            client_id=SHAREPOINT_CLIENT_ID,
            client_secret=SHAREPOINT_CLIENT_SECRET,
            site_url=SHAREPOINT_SITE_URL,
        )

        task["message"] = "Running SharePoint indexing pipeline …"
        summary = index_sharepoint(
            search_service=search_svc,
            vector_store=search_svc.vector_store,
            connector=connector,
            cache_dir=SHAREPOINT_CACHE_DIR,
            extract_fn=extract_text,
            make_chunks_fn=make_chunks,
            batch_size=BATCH_SIZE,
        )

        if summary.get("error"):
            task["status"]  = "error"
            task["message"] = summary["message"]
            logger.error("[task:%s] SharePoint indexing failed: %s", task_id, summary["message"])
        else:
            task["status"]  = "complete"
            task["message"] = summary["message"]
            task["summary"] = summary
            _connector_last_indexed["sharepoint"] = datetime.utcnow().isoformat()
            logger.info("[task:%s] SharePoint indexing complete: %s", task_id, summary)

    except Exception as exc:
        task["status"]  = "error"
        task["message"] = f"SharePoint indexing failed: {exc}"
        logger.exception("[task:%s] Unhandled error in SharePoint indexing", task_id)


@app.post("/api/index/sharepoint")
async def index_sharepoint_endpoint(background_tasks: BackgroundTasks):
    """
    Trigger a background SharePoint indexing run.

    - Returns 400 if ``SHAREPOINT_ENABLED`` is not true.
    - Returns 400 if any required credential env var is missing.
    - Otherwise queues the job and returns a ``task_id`` for polling via
      ``GET /api/crawl/status/{task_id}``.
    """
    from fastapi import HTTPException

    if not SHAREPOINT_ENABLED:
        raise HTTPException(
            status_code=400,
            detail=(
                "SharePoint connector is disabled. "
                "Set SHAREPOINT_ENABLED=true in .env to enable it."
            ),
        )

    missing = [v for v, val in [
        ("SHAREPOINT_TENANT_ID",     SHAREPOINT_TENANT_ID),
        ("SHAREPOINT_CLIENT_ID",     SHAREPOINT_CLIENT_ID),
        ("SHAREPOINT_CLIENT_SECRET", SHAREPOINT_CLIENT_SECRET),
        ("SHAREPOINT_SITE_URL",      SHAREPOINT_SITE_URL),
    ] if not val]

    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required env vars: {', '.join(missing)}",
        )

    task_id = str(uuid.uuid4())
    crawl_tasks[task_id] = {
        "status":       "queued",
        "message":      "Queued — waiting to start",
        "indexed":      0,
        "skipped":      0,
        "failed":       0,
        "total":        0,
        "total_chunks": 0,
    }
    background_tasks.add_task(background_index_sharepoint, task_id, _svc())
    logger.info("SharePoint indexing task queued: %s", task_id)
    return {"task_id": task_id, "status": "queued"}


# ── Search ─────────────────────────────────────────────────────────────────────

@app.get("/api/search")
def search(
    q: str = Query(..., min_length=2),
    sources: Optional[str] = Query(
        None,
        description="Comma-separated source filter: local,sharepoint,network_drive",
    ),
):
    """
    Search the vector index.

    Optional ``?sources=local,sharepoint`` filters results to only those
    connector sources.  Omit the param (or leave it empty) to search all.
    """
    source_list: Optional[list[str]] = None
    if sources:
        source_list = [s.strip() for s in sources.split(",") if s.strip()]
    return _svc().search(q, sources=source_list)


# ── Stats ──────────────────────────────────────────────────────────────────────

@app.get("/api/stats")
def stats():
    svc = _svc()
    if svc.vector_store is None:
        return {"total_chunks": 0, "total_files": 0, "file_types": {}, "model": MODEL_NAME, "mode": "no_index"}
    
    # Get stats from Qdrant
    qdrant_stats = svc.vector_store.get_stats()
    total_points = qdrant_stats.get("total_points", 0)
    
    if total_points == 0:
        return {"total_chunks": 0, "total_files": 0, "file_types": {}, "model": MODEL_NAME, "mode": "no_index"}
    
    # To get file types and unique files, we need to query Qdrant
    # For now, we'll do a simple scroll to get all points (this could be optimized for large datasets)
    try:
        from qdrant_client.models import ScrollRequest
        scroll_result = svc.vector_store.client.scroll(
            collection_name=svc.vector_store.collection_name,
            limit=10000,  # Adjust based on expected dataset size
            with_payload=True,
            with_vectors=False
        )
        
        file_types: dict[str, int] = {}
        unique_files: set[str] = set()
        
        for point in scroll_result[0]:
            payload = point.payload
            ext = payload.get("file_type", "?")
            file_types[ext] = file_types.get(ext, 0) + 1
            unique_files.add(payload.get("file_path", ""))
        
        return {
            "total_chunks": total_points,
            "total_files": len(unique_files),
            "file_types": file_types,
            "model": MODEL_NAME,
            "mode": "live"
        }
    except Exception as e:
        logger.error("Failed to get detailed stats: %s", e)
        return {
            "total_chunks": total_points,
            "total_files": 0,
            "file_types": {},
            "model": MODEL_NAME,
            "mode": "live"
        }


@app.get("/api/reload")
def reload_db():
    svc = _svc()
    stats = svc.vector_store.get_stats() if svc.vector_store else {"total_points": 0}
    logger.info("Manual reload triggered")
    return {"status": "reloaded", "vectors": stats.get("total_points", 0)}


@app.get("/api/open")
def open_file(path: str):
    try:
        p = path if os.path.isabs(path) else os.path.abspath(path)
        if platform.system() == "Windows":
            subprocess.Popen(["explorer", "/select,", p])
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", "-R", p])
        else:
            subprocess.Popen(["xdg-open", os.path.dirname(p)])
        logger.info("Opened file location: %s", p)
        return {"status": "opened", "path": p}
    except Exception as e:
        logger.error("Failed to open file %s: %s", path, e)
        return {"status": "error", "message": str(e)}


# ── Schedule & Health ──────────────────────────────────────────────────────────

@app.get("/api/schedule")
def get_schedule():
    """
    Return next scheduled run times and last-run timestamps for each source.
    """
    return _agent().get_schedule_info()


@app.get("/api/health")
def health_check():
    """
    Comprehensive health check: Qdrant connectivity, model status, connector
    states, and next scheduled run.
    """
    svc = _svc()

    # Qdrant
    qdrant_status = "connected"
    try:
        svc.vector_store.get_stats()
    except Exception as exc:
        qdrant_status = f"error: {exc}"

    # Next scheduled run across all jobs
    schedule = _agent().get_schedule_info()
    next_runs = [
        j["next_run"]
        for j in schedule["jobs"].values()
        if j.get("next_run")
    ]
    next_run = min(next_runs) if next_runs else None

    return {
        "status": "ok" if qdrant_status == "connected" else "degraded",
        "qdrant": qdrant_status,
        "model":  "loaded" if svc.model_loaded else "not_loaded",
        "sources": {
            "local": {
                "enabled":      True,
                "last_indexed": _connector_last_indexed.get("local"),
            },
            "sharepoint": {
                "enabled":      SHAREPOINT_ENABLED,
                "last_indexed": _connector_last_indexed.get("sharepoint"),
            },
            "network_drive": {
                "enabled":      NETWORK_DRIVE_ENABLED,
                "last_indexed": _connector_last_indexed.get("network_drive"),
            },
        },
        "next_scheduled_run": next_run,
    }


# ── Serve frontend ─────────────────────────────────────────────────────────────

@app.get("/")
def root():
    index_path = os.path.join(os.path.dirname(__file__), "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return {"message": "KnowledgeOS API is running"}


# ── Main ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host=API_HOST, port=API_PORT, reload=False)
